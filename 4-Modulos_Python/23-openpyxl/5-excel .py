"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl
from pathlib import Path
from random import uniform

ROOT = Path(__file__).parent
NOVA_PLANILHA = ROOT / 'novos_dados.xlsx'


planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedidos = linha - 1
    planilha1.cell(linha, 1).value = numero_pedidos
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = f'R$ {preco}'

for linha in range(1, 16):
    planilha2.cell(
        linha, 1).value = f'Willian {linha} R$ {round(uniform(10, 100), 2)}'
    planilha2.cell(
        linha, 2).value = f'Suelen {linha} R$ {round(uniform(10, 100), 2)}'
    planilha2.cell(
        linha, 3).value = f'Rafael {linha} R$ {round(uniform(10, 100), 2)}'

planilha.save(NOVA_PLANILHA)
