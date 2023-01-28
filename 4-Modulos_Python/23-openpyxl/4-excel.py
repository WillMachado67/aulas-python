"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl
from pathlib import Path
from random import uniform

ROOT = Path(__file__).parent
CAMINHO_PEDIDOS = ROOT / 'pedidos.xlsx'
NOVA_PLANILHAS = ROOT / 'novos_pedidos.xlsx'

pedidos = openpyxl.load_workbook(filename=CAMINHO_PEDIDOS)
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

for linha in range(5, 16):
    numero_pedidos = linha - 1
    planilha1.cell(linha, 1).value = numero_pedidos
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100))
    planilha1.cell(linha, 3).value = preco

pedidos.save(NOVA_PLANILHAS)
